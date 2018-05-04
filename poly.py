"""
    Генератор словаря с обходом кортежа заготовленных переменных
    и инициализацией их значений с помощью форматированного
    пользовательского ввода
    main_dict = {k: str(input("Set {}: ".format(k)))
             for k in defect_vars.MAIN_DEFECT_FIELDS}
    for k, v in main_dict.items():
            print('{}\n{} -- {}'.format(defect_vars.STAR_DELIM, k, v))
"""

"""Механизм инициации классов значениями словарей

class allMyFields(object):
    # note: you cannot (and don't have to) use self here
    Field1 = None
    Field2 = None

    def __init__(self, dictionary):
        self.__dict__.update(dictionary)

q = { 'Field1' : 3000, 'Field2' : 6000, 'RandomField' : 5000 }
instance = allMyFields(q)

print instance.Field1      # => 3000
print instance.Field2      # => 6000
print instance.RandomField # => 5000
"""
"""
import os

cwd = os.getcwd()
file_lst = os.listdir('.')
print(cwd)

for i in file_lst:
    print(i)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook('src/sheet.xlsx')
sheet = wb.get_sheet_by_name('Лист1')

df = pd.DataFrame(sheet.values)

new_wb = load_workbook('src/empty_sheet.xlsx')
new_ws = wb.active

for r in dataframe_to_rows(df, index=True, header=True):
    new_ws.append(r)

print(df)
"""
