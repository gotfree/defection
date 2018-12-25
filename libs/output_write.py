"""Модуль функций отображения и записи информации в файл

В состав модуля входят библиотеки:
    [openpyxl] -- [Чтение/запись объектов excel]

В состав модуля входят функции:
    [create_file] -- [Создание нового файла]
    [write_poly_tables] -- [Интерфейс главного меню программы]
    [write_tables] -- [Функция проверки пользовательского ввода]
"""

import openpyxl
from ..settings import BASE_DIR


def create_file(arg1, arg2, arg3):
    """Функция создания нового файла

    Args:
        arg1 (tuple): кортеж для ячеек заголовка файла (CUSTOMER_DEFECT_FIELDS)
        arg2 (tuple): кортеж для ячеек заголовка файла (GOODS_DEFECT_FIELDS)
        arg3 (tuple): кортеж для ячеек заголовка файла (RESULT_DEFECT_FIELDS)
    """
    head_tuple = arg1 + arg2 + arg3

    dest_file = 'output/test.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active

    # "Заморозка" ячеек заголовка файла
    ws.freeze_panes = 'A2'

    next_col = 1
    for i in head_tuple:
        ws.cell(column=next_col, row=1, value=i)
        next_col += 1

    # Задание ширины колонкам
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 35
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 22

    wb.save(filename=dest_file)


def write_poly_tables(arg1, arg2):
    """Запись в файл информации о клиенте и нескольких товарах

    - Требуется рефакторинг/декомпозиция
    - Рассмотреть вариант декоратора функции write_tables

    Args:
        arg1 (dict): Словарь с данными о контрагенте
        arg2 (list): Список словарей (частный случай - одного словаря)
        с информацией о товаре от конкретного контрагента
    """
    dest_file = 'output/test.xlsx'
    none_list = [None, None, None]

    # Запись информации о первом товаре
    value_poly_list = list(arg1.values()) + list(arg2[0].values())
    wpb = openpyxl.load_workbook(dest_file)
    wps = wpb.active

    next_col = 1
    row_num = wps.max_row + 1
    for i in value_poly_list:
        wps.cell(column=next_col, row=row_num, value=i)
        next_col += 1

    # "вытолкнули" первый элемент списка
    arg2.pop(0)
    for x in arg2:
        temp_list = none_list + list(x.values())
        next_col = 1
        row_num = wps.max_row + 1
        for i in temp_list:
            wps.cell(column=next_col, row=row_num, value=i)
            next_col += 1

    print("Информация успешно записана")
    wpb.save(filename=dest_file)


def write_tables(arg1, arg2):
    """Запись в файл информации о клиенте и об одном товаре

    Args:
        arg1 (dict): Словарь с данными о контрагенте
        arg2 (list): Список словарей (частный случай - одного словаря)
        с информацией о товаре от конкретного контрагента
    """
    dest_file = 'output/test.xlsx'
    value_list = list(arg1.values()) + list(arg2[0].values())
    wb = openpyxl.load_workbook(dest_file)
    ws = wb.active

    next_col = 1
    row_num = ws.max_row + 1
    for i in value_list:
        ws.cell(column=next_col, row=row_num, value=i)
        next_col += 1

    print("Информация успешно записана")
    wb.save(filename=dest_file)


if __name__ == '__main__':
    print(BASE_DIR)
